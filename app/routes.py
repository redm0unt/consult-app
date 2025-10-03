import secrets
import string
from typing import Optional, Tuple

from flask import Blueprint, abort, current_app, flash, jsonify, redirect, render_template, request, url_for
from flask_login import current_user, login_required
from sqlalchemy import func, select
from sqlalchemy.exc import IntegrityError, SQLAlchemyError

from openpyxl import load_workbook

from .auth.policies import AccountPolicy, UsersPolicy
from .models import Teacher, User, db
from .repositories import get_repository

user_repository = get_repository('users')


def _parse_full_name(full_name: str) -> Tuple[str, str, Optional[str]]:
    parts = [part for part in full_name.split() if part]
    if len(parts) < 2:
        raise ValueError('Пожалуйста, укажите как минимум фамилию и имя')
    if len(parts) > 3:
        raise ValueError('Пожалуйста, укажите не более трёх частей имени (Фамилия Имя Отчество)')

    if len(parts) == 2:
        last_name, first_name = parts
        return first_name, last_name, None

    last_name, first_name, middle_name = parts
    return first_name, last_name, middle_name


def _generate_password(length: int = 12) -> str:
    alphabet = string.ascii_letters + string.digits
    if length < 8:
        length = 8
    return ''.join(secrets.choice(alphabet) for _ in range(length))


bp = Blueprint('main', __name__)

def get_pages():
    return [('Главная', 'main.index'), ('Мероприятия', 'main.index'), ('Учителя', 'main.teachers'), ('Здания', 'main.index')]

@bp.route('/')
def index():
    return render_template('base.html',
                           pages=get_pages())

@bp.route('/login')
def login_redirect():
    return redirect(url_for('auth.login'))


@bp.route('/teachers', methods=['GET', 'POST'])
@login_required
def teachers():
    users_policy = UsersPolicy(user_id=current_user.user_id)
    if not users_policy.get_page():
        abort(403)

    school = current_user.school
    search_query = (request.args.get('q') or '').strip()
    form_data = {}
    form_mode = 'create'
    form_teacher_id: Optional[int] = None
    can_manage_users = users_policy.create()
    show_create_form = False

    if request.method == 'POST':
        if not can_manage_users:
            abort(403)

        form_type = request.form.get('form_type') or 'create'

        if form_type == 'delete':
            teacher_id = request.form.get('teacher_id', type=int)
            if not teacher_id:
                flash('Не удалось определить учителя для удаления.', 'warning')
                return redirect(url_for('main.teachers'))

            if not users_policy.delete():
                abort(403)

            teacher = db.session.get(Teacher, teacher_id)
            if not teacher or not school or teacher.school_id != school.school_id:
                flash('Учитель не найден или не относится к вашей школе.', 'warning')
                return redirect(url_for('main.teachers'))

            try:
                user_repository.delete(teacher.user_id)
            except SQLAlchemyError:
                user_repository.rollback()
                current_app.logger.exception('Failed to delete teacher %s', teacher_id)
                flash('Не удалось удалить учителя. Попробуйте ещё раз позже.', 'danger')
            else:
                flash('Учитель успешно удалён.', 'success')
            return redirect(url_for('main.teachers'))

        full_name_value = (request.form.get('full_name') or '').strip()
        email_value = (request.form.get('email') or '').strip()
        password_value = request.form.get('password') or ''

        form_data = {
            'full_name': full_name_value,
            'email': email_value,
        }

        errors = []
        first_name = last_name = middle_name = None

        if not full_name_value:
            errors.append('Укажите ФИО учителя.')
        else:
            try:
                first_name, last_name, middle_name = _parse_full_name(full_name_value)
            except ValueError as exc:
                errors.append(str(exc))

        email_lower = email_value.lower()
        if not email_value:
            errors.append('Укажите электронную почту учителя.')
        elif ('@' not in email_lower or email_lower.startswith('@') or email_lower.endswith('@')
              or ' ' in email_value or '.' not in email_lower.split('@', 1)[-1]):
            errors.append('Введите корректный адрес электронной почты.')

        if not school:
            errors.append('Невозможно привязать учителя к школе. Обратитесь к администратору системы.')

        if form_type == 'create':
            if not password_value:
                errors.append('Укажите временный пароль для учителя (минимум 8 символов).')
            elif len(password_value) < 8:
                errors.append('Пароль должен содержать не менее 8 символов.')

        teacher: Optional[Teacher] = None

        if form_type == 'update':
            form_mode = 'update'
            teacher_id = request.form.get('teacher_id', type=int)
            form_teacher_id = teacher_id
            if not teacher_id:
                errors.append('Не удалось определить учителя для обновления.')
            elif not users_policy.edit():
                abort(403)
            else:
                teacher = db.session.get(Teacher, teacher_id)
                if not teacher or not school or teacher.school_id != school.school_id:
                    errors.append('Учитель не найден или не относится к вашей школе.')

        if not errors and email_value:
            stmt = select(User).where(func.lower(User.email) == email_lower)
            if form_type == 'update' and teacher:
                stmt = stmt.where(User.user_id != teacher.user_id)
            existing_user = db.session.execute(stmt).scalar_one_or_none()
            if existing_user:
                errors.append('Пользователь с такой электронной почтой уже существует.')

        if not errors:
            try:
                if form_type == 'create':
                    user_repository.create(
                        email=email_lower,
                        password=password_value,
                        first_name=first_name,
                        last_name=last_name,
                        middle_name=middle_name,
                        school_id=school.school_id if school else None,
                        role='teacher',
                    )
                elif form_type == 'update' and teacher:
                    user_repository.update(
                        user_id=teacher.user_id,
                        email=email_lower,
                        first_name=first_name,
                        last_name=last_name,
                        middle_name=middle_name,
                    )
            except IntegrityError:
                user_repository.rollback()
                errors.append('Пользователь с такой электронной почтой уже существует.')
            except SQLAlchemyError:
                user_repository.rollback()
                action_error_msg = 'обновить' if form_type == 'update' else 'создать'
                current_app.logger.exception('Failed to %s teacher', action_error_msg)
                errors.append('Не удалось сохранить данные учителя. Попробуйте ещё раз позже.')

        if errors:
            show_create_form = True
            for error in errors:
                flash(error, 'warning')
        else:
            message = 'Учитель успешно обновлён.' if form_type == 'update' else 'Учитель успешно добавлен.'
            flash(message, 'success')
            return redirect(url_for('main.teachers'))

    teachers_list = []

    if school:
        stmt = (
            select(Teacher)
            .where(Teacher.school_id == school.school_id)
            .order_by(Teacher.last_name.asc(), Teacher.first_name.asc(), Teacher.middle_name.asc())
        )

        if search_query:
            like_pattern = f"%{search_query.lower()}%"
            stmt = stmt.where(
                func.lower(Teacher.email).like(like_pattern)
                | func.lower(Teacher.first_name).like(like_pattern)
                | func.lower(Teacher.last_name).like(like_pattern)
                | func.lower(func.coalesce(Teacher.middle_name, '')).like(like_pattern)
            )

        teachers_list = db.session.execute(stmt).scalars().all()

    return render_template(
        'teachers/index.html',
        pages=get_pages(),
        teachers=teachers_list,
        search_query=search_query,
        can_manage_users=can_manage_users,
        teacher_form_data=form_data,
        show_teacher_form=show_create_form,
        teacher_form_mode=form_mode,
        teacher_form_teacher_id=form_teacher_id,
    )


@bp.route('/teachers/import', methods=['POST'])
@login_required
def teachers_import():
    users_policy = UsersPolicy(user_id=current_user.user_id)
    if not users_policy.create():
        abort(403)

    school = current_user.school
    if not school:
        return jsonify({
            'success': False,
            'message': 'Невозможно определить школу для привязки учителей. Обратитесь к администратору системы.',
        }), 400

    upload = request.files.get('file')
    if not upload or not upload.filename:
        return jsonify({
            'success': False,
            'message': 'Выберите Excel-файл с учителями перед загрузкой.',
        }), 400

    filename_lower = upload.filename.lower()
    if not filename_lower.endswith('.xlsx'):
        return jsonify({
            'success': False,
            'message': 'Поддерживается только импорт файлов в формате .xlsx.',
        }), 400

    try:
        workbook = load_workbook(upload, data_only=True)
    except Exception:  # pragma: no cover - handled by error response
        current_app.logger.exception('Failed to load teachers workbook')
        return jsonify({
            'success': False,
            'message': 'Не удалось прочитать файл. Убедитесь, что это корректный Excel-документ (.xlsx).',
        }), 400

    try:
        worksheet = workbook.active
        rows = [tuple(row) for row in worksheet.iter_rows(values_only=True)]
    finally:
        workbook.close()

    if not rows:
        return jsonify({
            'success': False,
            'message': 'Файл пуст. Добавьте данные об учителях и повторите попытку.',
        }), 400

    header_row = rows[0]
    normalized_header = [str(cell).strip().lower() if isinstance(cell, str) else '' for cell in header_row]
    has_header = 'email' in normalized_header and 'full_name' in normalized_header

    if has_header:
        email_index = normalized_header.index('email')
        fullname_index = normalized_header.index('full_name')
        data_rows = rows[1:]
        row_start_index = 2
    else:
        email_index = 0
        fullname_index = 1
        data_rows = rows
        row_start_index = 1

    if email_index == fullname_index:
        return jsonify({
            'success': False,
            'message': 'Файл должен содержать отдельные колонки для email и ФИО учителя.',
        }), 400

    created_teachers: list[dict[str, str]] = []
    errors: list[str] = []
    seen_emails: set[str] = set()

    for offset, row in enumerate(data_rows, start=row_start_index):
        row_values = list(row)
        email_raw = row_values[email_index] if len(row_values) > email_index else None
        full_name_raw = row_values[fullname_index] if len(row_values) > fullname_index else None

        if email_raw is None and full_name_raw is None:
            continue

        email_value = str(email_raw).strip() if email_raw is not None else ''
        full_name_value = str(full_name_raw).strip() if full_name_raw is not None else ''

        if not email_value:
            errors.append(f'Строка {offset}: не указана электронная почта учителя.')
            continue

        email_lower = email_value.lower()
        if ('@' not in email_lower or email_lower.startswith('@') or email_lower.endswith('@')
                or ' ' in email_value or '.' not in email_lower.split('@', 1)[-1]):
            errors.append(f'Строка {offset}: некорректный адрес электронной почты — "{email_value}".')
            continue

        if email_lower in seen_emails:
            errors.append(f'Строка {offset}: адрес "{email_value}" уже встречался в файле. Дубликат пропущен.')
            continue

        seen_emails.add(email_lower)

        if not full_name_value:
            errors.append(f'Строка {offset}: не указано ФИО учителя.')
            continue

        try:
            first_name, last_name, middle_name = _parse_full_name(full_name_value)
        except ValueError as exc:
            errors.append(f'Строка {offset}: {exc}')
            continue

        stmt = select(User).where(func.lower(User.email) == email_lower)
        existing_user = db.session.execute(stmt).scalar_one_or_none()
        if existing_user:
            errors.append(f'Строка {offset}: пользователь с электронной почтой "{email_value}" уже существует.')
            continue

        password = _generate_password()

        try:
            user_repository.create(
                email=email_lower,
                password=password,
                first_name=first_name,
                last_name=last_name,
                middle_name=middle_name,
                school_id=school.school_id,
                role='teacher',
            )
        except IntegrityError:
            user_repository.rollback()
            errors.append(f'Строка {offset}: пользователь с электронной почтой "{email_value}" уже существует.')
            continue
        except SQLAlchemyError:
            user_repository.rollback()
            current_app.logger.exception('Failed to create teacher from import (row %s)', offset)
            errors.append(f'Строка {offset}: не удалось создать учителя из-за ошибки базы данных.')
            continue

        recorded_full_name = ' '.join(filter(None, [last_name, first_name, middle_name]))
        created_teachers.append({
            'full_name': recorded_full_name,
            'email': email_lower,
            'password': password,
        })

    if not created_teachers:
        return jsonify({
            'success': False,
            'message': 'Учителя не были созданы.',
            'errors': errors or ['Файл не содержит корректных данных для импорта.'],
        }), 400

    return jsonify({
        'success': True,
        'message': f'Добавлено учителей: {len(created_teachers)}.',
        'count': len(created_teachers),
        'errors': errors,
        'teachers': created_teachers,
    })


@bp.route('/account', methods=['GET', 'POST'])
@login_required
def account():
    policy = AccountPolicy(target_user_id=current_user.user_id)

    if not policy.show():
        abort(403)

    school = current_user.school
    can_manage_school = policy.manage_school()

    if request.method == 'POST':
        if not policy.edit():
            abort(403)

        form = request.form
        errors = []
        updated = False

        email_input = (form.get('email') or '').strip()
        normalized_email = email_input.lower()

        first_name = (form.get('first_name') or '').strip()
        last_name = (form.get('last_name') or '').strip()
        middle_name = (form.get('middle_name') or '').strip()
        school_name = (form.get('school_name') or '').strip()
        password_old = form.get('password_old') or ''
        password_new = form.get('password_new') or ''

        if not email_input:
            errors.append('Электронная почта не может быть пустой.')
        elif ('@' not in normalized_email or normalized_email.startswith('@') or normalized_email.endswith('@') or
              ' ' in email_input or '.' not in normalized_email.split('@', 1)[-1]):
            errors.append('Введите корректный адрес электронной почты.')
        else:
            current_email_normalized = (current_user.email or '').lower()
            if normalized_email != current_email_normalized:
                stmt = select(User).where(User.email == normalized_email)
                existing_user = db.session.execute(stmt).scalar_one_or_none()
                if existing_user and existing_user.user_id != current_user.user_id:
                    errors.append('Пользователь с такой электронной почтой уже существует.')
                else:
                    current_user.email = normalized_email
                    updated = True

        if not first_name:
            errors.append('Имя не может быть пустым.')
        elif first_name != current_user.first_name:
            current_user.first_name = first_name
            updated = True

        if not last_name:
            errors.append('Фамилия не может быть пустой.')
        elif last_name != current_user.last_name:
            current_user.last_name = last_name
            updated = True

        middle_value = middle_name or None
        if current_user.middle_name != middle_value:
            current_user.middle_name = middle_value
            updated = True

        if can_manage_school:
            if not school:
                if school_name:
                    errors.append('Не удалось найти школу для обновления. Обратитесь к администратору.')
            else:
                if not school_name:
                    errors.append('Название школы не может быть пустым.')
                elif school.school_name != school_name:
                    school.school_name = school_name
                    updated = True

        if password_old or password_new:
            if not password_old or not password_new:
                errors.append('Для смены пароля заполните оба поля.')
            elif not current_user.check_password(password_old):
                errors.append('Старый пароль указан неверно.')
            elif len(password_new) < 8:
                errors.append('Новый пароль должен содержать не менее 8 символов.')
            else:
                current_user.set_password(password_new)
                updated = True

        if errors:
            db.session.rollback()
            for message in errors:
                flash(message, 'warning')
        elif updated:
            try:
                db.session.commit()
            except SQLAlchemyError:
                db.session.rollback()
                current_app.logger.exception('Failed to update account information')
                flash('Не удалось сохранить изменения. Попробуйте ещё раз позже.', 'danger')
            else:
                flash('Изменения успешно сохранены.', 'success')
                return redirect(url_for('main.account'))
        else:
            flash('Изменений не обнаружено.', 'info')
            return redirect(url_for('main.account'))

        school = current_user.school

    invite_link = None

    if policy.view_invite_link() and school and school.invite_code:
        invite_link = url_for('auth.register_parent', invite_code=school.invite_code, _external=True)

    return render_template(
        'account/account.html',
        pages=get_pages(),
        school=school,
        invite_link=invite_link,
        can_manage_school=can_manage_school,
    )
